import openpyxl
from tabulate import tabulate


excel = openpyxl.load_workbook("personas.xlsx")


# Creamos una variable para sacar le libro que está activo
libro_activo = excel.active


# Creamos una lista vacía para almacenar los datos
datos = []

# Usamos max_row para que recorra todas las filas del archivo
for row in range(1, libro_activo.max_row):
    # Creamos una lista vacía para almacenar los datos de cada fila del archivo excel
    _row = [
        row,
    ]

    # Usamos iter_cols para que recorra las columnas
    for col in libro_activo.iter_cols(1, libro_activo.max_column):
        # Agregamos los datos de cada columna a la lista _row
        _row.append(col[row].value)

    # Agregamos los datos de cada fila a la lista datos
    datos.append(_row)

# Creamos una lista para almacenar los encabezados
headers = ["#", "ID", "Nombre", "Company", "Email", "Mac_Adress"]

header_align = ("center",) * len(headers)


# Usamos tablefmt para darle formato a la tabla (fancy_grid) = Tabla con bordes
print(tabulate(datos, headers=headers, tablefmt="fancy_grid", colalign=header_align))
